"""
Excel 리포트 자동 생성 유틸리티
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from datetime import datetime, timedelta
import os

def generate_excel_report(db, analyzer, months=6, include_charts=True):
    """Excel 리포트 생성"""
    
    wb = Workbook()
    
    # 기본 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. 요약 시트
    create_summary_sheet(wb, db, analyzer, months)
    
    # 2. 월별 상세 시트
    create_monthly_detail_sheet(wb, db, analyzer, months)
    
    # 3. 카테고리 분석 시트
    create_category_analysis_sheet(wb, db, analyzer, months)
    
    # 4. 거래 내역 시트
    create_transactions_sheet(wb, db, months)
    
    # 5. 목표 추적 시트
    create_goals_sheet(wb, db, analyzer)
    
    # 파일 저장
    output_dir = '/home/claude/card-expense-manager/data'
    os.makedirs(output_dir, exist_ok=True)
    
    filename = f'지출분석_리포트_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    return filepath

def create_summary_sheet(wb, db, analyzer, months):
    """요약 시트 생성"""
    ws = wb.create_sheet('요약', 0)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=14)
    
    # 제목
    ws['A1'] = '카드 지출 분석 리포트'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'생성일: {datetime.now().strftime("%Y년 %m월 %d일")}'
    ws.merge_cells('A2:D2')
    
    # 기간별 지출 요약
    ws['A4'] = '기간별 지출 요약'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws.merge_cells('A4:D4')
    
    ws['A5'] = '기간'
    ws['B5'] = '총 지출'
    ws['C5'] = '거래 건수'
    ws['D5'] = '일평균'
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}5'].font = Font(bold=True)
        ws[f'{col}5'].fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    
    # 월별 데이터
    row = 6
    now = datetime.now()
    
    for i in range(months - 1, -1, -1):
        target_date = now - timedelta(days=30 * i)
        year = target_date.year
        month = target_date.month
        
        report = analyzer.get_monthly_report(year, month)
        
        ws[f'A{row}'] = f'{year}년 {month}월'
        ws[f'B{row}'] = report['total_expense']
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = report['transaction_count']
        ws[f'D{row}'] = report.get('avg_daily_expense', 0)
        ws[f'D{row}'].number_format = '#,##0'
        
        row += 1
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15

def create_monthly_detail_sheet(wb, db, analyzer, months):
    """월별 상세 시트"""
    ws = wb.create_sheet('월별 상세')
    
    # 헤더
    headers = ['월', '카테고리', '지출액', '거래건수', '평균']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    row = 2
    now = datetime.now()
    
    for i in range(months - 1, -1, -1):
        target_date = now - timedelta(days=30 * i)
        year = target_date.year
        month = target_date.month
        
        report = analyzer.get_monthly_report(year, month)
        
        for category, data in sorted(
            report['by_category'].items(),
            key=lambda x: x[1]['total'],
            reverse=True
        ):
            ws.cell(row, 1, f'{year}-{month:02d}')
            ws.cell(row, 2, category)
            ws.cell(row, 3, data['total'])
            ws.cell(row, 3).number_format = '#,##0'
            ws.cell(row, 4, data['count'])
            ws.cell(row, 5, data['total'] // data['count'] if data['count'] > 0 else 0)
            ws.cell(row, 5).number_format = '#,##0'
            
            row += 1
    
    # 열 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

def create_category_analysis_sheet(wb, db, analyzer, months):
    """카테고리 분석 시트"""
    ws = wb.create_sheet('카테고리 분석')
    
    # 전체 기간 통계
    stats = analyzer.get_category_stats()
    
    # 헤더
    headers = ['카테고리', '총 지출', '거래건수', '비율(%)', '평균 금액']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    
    # 데이터
    for row, stat in enumerate(stats, 2):
        ws.cell(row, 1, stat['category'])
        ws.cell(row, 2, stat['total'])
        ws.cell(row, 2).number_format = '#,##0'
        ws.cell(row, 3, stat['count'])
        ws.cell(row, 4, stat['percentage'])
        ws.cell(row, 4).number_format = '0.0'
        ws.cell(row, 5, stat['avg_per_transaction'])
        ws.cell(row, 5).number_format = '#,##0'
    
    # 차트 추가
    chart = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(stats) + 1)
    data = Reference(ws, min_col=2, min_row=1, max_row=len(stats) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "카테고리별 지출 비율"
    
    ws.add_chart(chart, "G2")
    
    # 열 너비
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

def create_transactions_sheet(wb, db, months):
    """거래 내역 시트"""
    ws = wb.create_sheet('거래내역')
    
    # 헤더
    headers = ['날짜', '가맹점', '카테고리', '카드', '금액']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    
    # 데이터 (최근 1000건)
    transactions = db.get_transactions(limit=1000)
    
    for row, trans in enumerate(transactions, 2):
        ws.cell(row, 1, trans['date'])
        ws.cell(row, 2, trans['merchant'])
        ws.cell(row, 3, trans['category'])
        ws.cell(row, 4, trans['card_name'])
        ws.cell(row, 5, trans['amount'])
        ws.cell(row, 5).number_format = '#,##0'
    
    # 열 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

def create_goals_sheet(wb, db, analyzer):
    """목표 추적 시트"""
    ws = wb.create_sheet('재무목표')
    
    goals = db.get_goals()
    
    if not goals:
        ws['A1'] = '설정된 재무 목표가 없습니다.'
        return
    
    # 헤더
    headers = ['목표명', '목표금액', '현재금액', '달성률(%)', '목표일자', '남은기간(월)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    
    for row, goal in enumerate(goals, 2):
        progress = analyzer.calculate_goal_progress(goal)
        
        ws.cell(row, 1, goal['name'])
        ws.cell(row, 2, goal['target_amount'])
        ws.cell(row, 2).number_format = '#,##0'
        ws.cell(row, 3, progress['current_amount'])
        ws.cell(row, 3).number_format = '#,##0'
        ws.cell(row, 4, progress['percentage'])
        ws.cell(row, 4).number_format = '0.0'
        ws.cell(row, 5, goal['target_date'])
        ws.cell(row, 6, progress.get('remaining_months', 0))
    
    # 열 너비
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
